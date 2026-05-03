"""
Generate testcases.xlsx — danh sách test case toàn hệ thống e-commerce.

Cột:
  STT | Module | Màn hình | Service | Điều kiện tiền đề | Thao tác thực hiện
      | Kết quả mong đợi | Người test | Ngày test | Kết quả

- Cột "Kết quả": data validation dropdown {OK, NG, NA, ""}.
- Conditional formatting: row đổi màu theo trạng thái (OK=xanh, NG=đỏ, NA=xám).
- Header bold, freeze panes, autofilter, column width auto.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# TEST CASE DATA
# ============================================================
# Mỗi dict: module, screen, service, precondition, steps, expected
# STT tự đánh; người test/ngày test/kết quả để trống cho tester điền.
# ============================================================

TESTCASES = [
    # ===== AUTHENTICATION =====
    dict(module="Authentication", screen="/register", service="user-service POST /api/users/auth/register",
         precondition="Email chưa đăng ký trong DB",
         steps="1. Truy cập /register\n2. Nhập tên, email mới, mật khẩu hợp lệ (>=8 ký tự)\n3. Nhấn 'Đăng ký'",
         expected="Tài khoản tạo thành công, auto login, redirect về /, Header hiển thị tên user"),
    dict(module="Authentication", screen="/register", service="user-service POST /api/users/auth/register",
         precondition="Email ABC@x.com đã tồn tại",
         steps="1. Truy cập /register\n2. Nhập email ABC@x.com\n3. Nhấn 'Đăng ký'",
         expected="Banner lỗi 'Email đã được sử dụng', không tạo account, ở lại /register"),
    dict(module="Authentication", screen="/register", service="frontend validation",
         precondition="Không có",
         steps="1. /register\n2. Nhập email không hợp lệ 'abc'\n3. Nhấn 'Đăng ký'",
         expected="Field email báo 'Email không hợp lệ', form không submit"),
    dict(module="Authentication", screen="/register", service="frontend validation",
         precondition="Không có",
         steps="1. /register\n2. Nhập password '123' (<8 ký tự)\n3. Nhấn 'Đăng ký'",
         expected="Field password báo lỗi độ dài tối thiểu, form không submit"),
    dict(module="Authentication", screen="/login", service="user-service POST /api/users/auth/login",
         precondition="Tài khoản user@x.com / 12345678 đã tồn tại",
         steps="1. /login\n2. Nhập user@x.com / 12345678\n3. Nhấn 'Đăng nhập'",
         expected="Login thành công, localStorage có accessToken + userProfile, redirect về / (hoặc returnTo nếu có)"),
    dict(module="Authentication", screen="/login", service="user-service POST /api/users/auth/login",
         precondition="Tài khoản tồn tại",
         steps="1. /login\n2. Nhập email đúng + mật khẩu sai\n3. Nhấn 'Đăng nhập'",
         expected="Banner 'Email hoặc mật khẩu không chính xác', KHÔNG redirect, KHÔNG xóa token cũ"),
    dict(module="Authentication", screen="/login?returnTo=/cart", service="frontend redirect",
         precondition="Đã có tài khoản hợp lệ",
         steps="1. /login?returnTo=/cart\n2. Đăng nhập đúng",
         expected="Redirect về /cart sau khi auth thành công"),
    dict(module="Authentication", screen="/login?returnTo=/login", service="frontend sanitizeReturnTo",
         precondition="Đã có tài khoản hợp lệ",
         steps="1. /login?returnTo=/login\n2. Đăng nhập đúng",
         expected="Sanitize returnTo='/login' thành '/' → redirect về trang chủ, KHÔNG vòng lặp"),
    dict(module="Authentication", screen="Header", service="frontend logout",
         precondition="Đã đăng nhập",
         steps="1. Click avatar/menu user trên Header\n2. Chọn 'Đăng xuất'",
         expected="Token + userProfile bị xóa, cart query cache xóa, redirect về /, Header chuyển về 'Đăng nhập/Đăng ký'"),
    dict(module="Authentication", screen="Cross-tab", service="frontend storage event",
         precondition="2 tab cùng user đăng nhập",
         steps="1. Tab A: Đăng xuất\n2. Quay sang Tab B",
         expected="Tab B tự cập nhật trạng thái không đăng nhập (không cần reload)"),

    # ===== PRODUCT BROWSING =====
    dict(module="Product", screen="/", service="product-service GET /api/products",
         precondition="DB có ≥10 sản phẩm active",
         steps="1. Mở /\n2. Cuộn xuống section sản phẩm",
         expected="Hiển thị grid sản phẩm có ảnh, tên, giá, badge stock; không vỡ layout"),
    dict(module="Product", screen="/products", service="product-service GET /api/products",
         precondition="DB có nhiều sản phẩm và categories",
         steps="1. /products\n2. Lọc theo category 'Phở'",
         expected="Chỉ hiển thị sản phẩm thuộc category 'Phở'; URL có query param category"),
    dict(module="Product", screen="/products", service="product-service GET /api/products?q=...",
         precondition="DB có sản phẩm tên chứa 'phở'",
         steps="1. /products\n2. Nhập 'phở' vào ô tìm kiếm\n3. Submit",
         expected="Danh sách lọc theo keyword; phân trang chính xác"),
    dict(module="Product", screen="/products/[slug]", service="product-service GET /api/products/slug/{slug}",
         precondition="Sản phẩm slug 'pho-bo-nam-dinh' tồn tại, stock>0",
         steps="1. Click vào sản phẩm trong grid\n2. Xem trang chi tiết",
         expected="Hiển thị tên, mô tả, ảnh chính + thumbnail, giá, badge tồn kho ('Còn nhiều'/'Sắp hết'/'Hết hàng'), nút Thêm vào giỏ"),
    dict(module="Product", screen="/products/[slug]", service="frontend stock display",
         precondition="Sản phẩm có stock = 0",
         steps="1. Mở trang chi tiết sản phẩm hết hàng",
         expected="Hiển thị badge 'Hết hàng', KHÔNG render nút 'Thêm vào giỏ' và quantity selector"),
    dict(module="Product", screen="/products/[slug]", service="frontend not found",
         precondition="Slug không tồn tại trong DB",
         steps="1. /products/khong-ton-tai",
         expected="Trang 404 (Next.js notFound) hiển thị thông báo phù hợp"),

    # ===== CART =====
    dict(module="Cart", screen="/products/[slug]", service="order-service POST /api/orders/cart/items",
         precondition="Đã đăng nhập, sản phẩm còn stock ≥ 5",
         steps="1. Trang chi tiết sản phẩm\n2. Chọn quantity = 2\n3. Click 'Thêm vào giỏ'",
         expected="Toast 'Đã thêm vào giỏ hàng' (success), Header cart count tăng đúng số lượng, DB cart_items có bản ghi mới"),
    dict(module="Cart", screen="/products/[slug]", service="order-service stock validate",
         precondition="Đã đăng nhập, sản phẩm có stock = 0 (DB)",
         steps="1. Trang chi tiết SP hết hàng (nếu nút bị ẩn → đổi qty manual hoặc backend)\n2. Click 'Thêm vào giỏ'",
         expected="Toast lỗi 'Sản phẩm không đủ tồn kho (... : còn 0)', cart KHÔNG thay đổi"),
    dict(module="Cart", screen="/cart", service="order-service GET /api/orders/cart",
         precondition="Đã đăng nhập, cart có items",
         steps="1. Click icon cart trên Header → /cart",
         expected="Hiển thị danh sách items với ảnh, tên, giá, số lượng, tổng tiền, nút +/− và X (xóa)"),
    dict(module="Cart", screen="/cart", service="order-service PATCH /api/orders/cart/items/{id}",
         precondition="Cart có item qty=1, stock backend=10",
         steps="1. /cart\n2. Click '+' để tăng quantity",
         expected="Quantity tăng 1, tổng tiền recompute, request PATCH thành công"),
    dict(module="Cart", screen="/cart", service="order-service PATCH stock validate",
         precondition="Cart có item qty đã đạt stock max (vd 5)",
         steps="1. /cart\n2. Click '+' khi qty đã = stock",
         expected="Nút '+' disabled hoặc toast 'Số lượng vượt quá tồn kho', qty KHÔNG vượt"),
    dict(module="Cart", screen="/cart", service="order-service PATCH /api/orders/cart/items/{id}",
         precondition="Cart có item qty=2",
         steps="1. /cart\n2. Click '−' giảm về 1",
         expected="Qty giảm còn 1, tổng tiền recompute"),
    dict(module="Cart", screen="/cart", service="order-service DELETE /api/orders/cart/items/{id}",
         precondition="Cart có ≥1 item",
         steps="1. /cart\n2. Click X trên 1 item",
         expected="Item bị xóa khỏi list (hoặc confirm dialog), DB cart_items mất bản ghi"),
    dict(module="Cart", screen="/cart", service="frontend empty state",
         precondition="Cart rỗng (đã đăng nhập)",
         steps="1. /cart",
         expected="Hiển thị empty state 'Giỏ hàng trống' + nút 'Tiếp tục mua sắm' về /products"),
    dict(module="Cart", screen="/cart", service="frontend auth guard",
         precondition="Chưa đăng nhập",
         steps="1. Truy cập trực tiếp /cart",
         expected="Redirect về /login?returnTo=/cart hoặc hiển thị prompt đăng nhập"),
    dict(module="Cart", screen="login flow", service="order-service POST /api/orders/cart/merge",
         precondition="Guest có cart trong localStorage; backend cart user rỗng",
         steps="1. Đăng nhập\n2. Vào /cart sau khi login",
         expected="Items từ guest cart được merge vào server cart, localStorage.cart bị clear; nếu merge fail → toast warning, login vẫn thành công"),

    # ===== CHECKOUT =====
    dict(module="Checkout", screen="/checkout", service="frontend route guard",
         precondition="Cart rỗng",
         steps="1. Truy cập /checkout",
         expected="Redirect về /cart hoặc hiển thị thông báo cart rỗng"),
    dict(module="Checkout", screen="/checkout", service="user-service GET /api/users/addresses",
         precondition="Cart có items, user chưa có địa chỉ",
         steps="1. /checkout",
         expected="Hiển thị form thêm địa chỉ mới hoặc CTA 'Thêm địa chỉ giao hàng'"),
    dict(module="Checkout", screen="/checkout", service="order-service POST /api/orders",
         precondition="Cart có items, đã có địa chỉ default",
         steps="1. /checkout\n2. Chọn địa chỉ giao hàng\n3. Chọn payment method 'COD'\n4. Nhấn 'Đặt hàng'",
         expected="Order tạo thành công (status PENDING), redirect về /profile/orders/{id}, cart bị clear, stock decrement đúng"),
    dict(module="Checkout", screen="/checkout", service="order-service stock revalidate",
         precondition="Cart có item qty=5, nhưng admin vừa giảm stock SP đó về 2",
         steps="1. /checkout\n2. Đặt hàng",
         expected="Backend trả 409 STOCK_SHORTAGE với detail 'còn 2', toast lỗi rõ ràng, order KHÔNG tạo"),
    dict(module="Checkout", screen="/checkout", service="order-service coupon redeem",
         precondition="Cart có items, có coupon 'SALE10' active",
         steps="1. /checkout\n2. Nhập 'SALE10'\n3. Click 'Áp dụng'\n4. Đặt hàng",
         expected="Tổng tiền giảm theo coupon, order ghi couponCode + discountAmount, coupon usage tăng +1 trong DB"),
    dict(module="Checkout", screen="/checkout", service="order-service coupon validation",
         precondition="Coupon đã hết hạn / đã đạt usage limit",
         steps="1. /checkout\n2. Nhập coupon hết hạn\n3. Áp dụng",
         expected="Toast lỗi 'Mã không hợp lệ' / 'Đã hết lượt sử dụng', tổng tiền không đổi"),
    dict(module="Checkout", screen="/checkout", service="order-service coupon validation",
         precondition="Coupon yêu cầu min order ≥ 500k, cart 100k",
         steps="1. Nhập coupon\n2. Áp dụng",
         expected="Toast lỗi 'Đơn hàng chưa đạt giá trị tối thiểu', không apply"),

    # ===== ORDERS / PROFILE =====
    dict(module="Orders", screen="/profile/orders", service="order-service GET /api/orders",
         precondition="User A đã đặt 3 đơn",
         steps="1. /profile (hoặc /profile/orders)\n2. Xem danh sách đơn hàng",
         expected="Chỉ hiển thị 3 đơn của user A, sắp xếp theo updatedAt desc"),
    dict(module="Orders", screen="/profile/orders", service="security cross-user",
         precondition="User A có đơn id=ABC, đăng nhập user B (chưa có đơn)",
         steps="1. Login user B\n2. Xem /profile/orders",
         expected="Danh sách rỗng, KHÔNG thấy đơn của user A"),
    dict(module="Orders", screen="/profile/orders", service="order-service filter status",
         precondition="User có nhiều đơn nhiều trạng thái",
         steps="1. /profile/orders\n2. Lọc 'PENDING'",
         expected="Chỉ hiển thị đơn PENDING; URL có query ?status=PENDING"),
    dict(module="Orders", screen="/profile/orders", service="order-service filter date",
         precondition="User có đơn cả tháng trước và tháng này",
         steps="1. /profile/orders\n2. Chọn from/to trong tháng này",
         expected="Chỉ đơn trong khoảng date hiển thị"),
    dict(module="Orders", screen="/profile/orders/[id]", service="order-service GET /api/orders/{id}",
         precondition="User có đơn id=XYZ",
         steps="1. /profile/orders\n2. Click vào đơn XYZ",
         expected="Trang chi tiết đơn hiển thị items, status tracker, địa chỉ, payment method, tổng tiền chính xác"),
    dict(module="Orders", screen="/profile/orders/[id]", service="security IDOR",
         precondition="User A có đơn id=ABC, đăng nhập user B",
         steps="1. Login user B\n2. Truy cập /profile/orders/ABC",
         expected="Hiển thị 'Không tìm thấy đơn hàng' (404), KHÔNG load được data của user A"),
    dict(module="Orders", screen="/profile/orders/[id]", service="frontend useParams",
         precondition="Đơn hàng id=XYZ tồn tại",
         steps="1. Mở trực tiếp URL /profile/orders/XYZ",
         expected="Page resolve param đúng, fetch /api/orders/XYZ (không phải /api/orders/undefined)"),

    # ===== PROFILE / ADDRESSES =====
    dict(module="Profile", screen="/profile", service="user-service GET /api/users/me",
         precondition="Đã đăng nhập",
         steps="1. Click avatar → /profile",
         expected="Hiển thị thông tin cá nhân (tên, email, phone), tabs Đơn hàng / Địa chỉ / Cài đặt"),
    dict(module="Profile", screen="/profile/settings", service="user-service PATCH /api/users/me",
         precondition="Đã đăng nhập",
         steps="1. /profile/settings\n2. Đổi tên hiển thị\n3. Lưu",
         expected="Toast success, tên cập nhật trong Header và localStorage userProfile"),
    dict(module="Profile", screen="/profile/addresses", service="user-service GET /api/users/addresses",
         precondition="User chưa có địa chỉ",
         steps="1. /profile/addresses",
         expected="Empty state + nút 'Thêm địa chỉ mới'"),
    dict(module="Profile", screen="/profile/addresses", service="user-service POST /api/users/addresses",
         precondition="Đã đăng nhập",
         steps="1. /profile/addresses\n2. Thêm địa chỉ\n3. Điền street/ward/district/city/phone\n4. Lưu",
         expected="Địa chỉ mới hiển thị trong list, đánh dấu default nếu là địa chỉ đầu tiên"),
    dict(module="Profile", screen="/profile/addresses", service="user-service PATCH default",
         precondition="User có ≥2 địa chỉ",
         steps="1. /profile/addresses\n2. Click 'Đặt làm mặc định' trên địa chỉ chưa default",
         expected="Địa chỉ được chọn trở thành default, cũ hủy default"),
    dict(module="Profile", screen="/profile/addresses", service="user-service DELETE",
         precondition="User có ≥1 địa chỉ",
         steps="1. /profile/addresses\n2. Click X trên địa chỉ\n3. Confirm",
         expected="Địa chỉ bị xóa, list cập nhật"),

    # ===== REVIEWS =====
    dict(module="Reviews", screen="/products/[slug] - tab Reviews", service="review-service GET reviews",
         precondition="Sản phẩm có ≥3 reviews",
         steps="1. Trang chi tiết SP\n2. Click tab 'Đánh giá'",
         expected="Hiển thị danh sách reviews (rating, text, tên, ngày), avg rating, phân phối stars"),
    dict(module="Reviews", screen="/products/[slug] - tab Reviews", service="review-service POST review",
         precondition="Đã đăng nhập, đã mua + nhận hàng SP đó",
         steps="1. Tab 'Đánh giá'\n2. Click 'Viết đánh giá'\n3. Chọn 5 sao + nhập text\n4. Submit",
         expected="Review xuất hiện trong list, avg rating cập nhật, badge 'Đã mua'"),
    dict(module="Reviews", screen="/products/[slug]", service="review-service auth guard",
         precondition="Chưa đăng nhập",
         steps="1. Tab 'Đánh giá'\n2. Cố gắng viết review",
         expected="Prompt đăng nhập hoặc nút bị ẩn"),
    dict(module="Reviews", screen="/products/[slug]", service="review-service eligibility",
         precondition="Đã đăng nhập nhưng CHƯA mua SP",
         steps="1. Tab 'Đánh giá'\n2. Cố gắng viết review",
         expected="Hiển thị 'Bạn cần mua sản phẩm để đánh giá', nút ẩn hoặc disabled"),

    # ===== ADMIN =====
    dict(module="Admin", screen="/admin", service="frontend role guard",
         precondition="Đăng nhập với role USER (không phải ADMIN)",
         steps="1. Truy cập /admin",
         expected="Redirect về / hoặc 403, không thấy menu admin"),
    dict(module="Admin", screen="/admin", service="frontend role guard",
         precondition="Đăng nhập với role ADMIN",
         steps="1. /admin",
         expected="Hiển thị dashboard admin với chart doanh thu, low-stock, top products"),
    dict(module="Admin", screen="/admin/products", service="product-service admin CRUD",
         precondition="Admin đăng nhập, DB có products",
         steps="1. /admin/products\n2. Xem danh sách",
         expected="Bảng products với cột id, name, price, stock, category, action; pagination + search"),
    dict(module="Admin", screen="/admin/products", service="product-service POST /api/products/admin",
         precondition="Admin đăng nhập",
         steps="1. /admin/products\n2. Click 'Thêm SP'\n3. Điền form (name, price, stock, brand, category, mô tả, ảnh URL)\n4. Lưu",
         expected="SP mới xuất hiện trong list, slug auto-generate, có thể view ở /products"),
    dict(module="Admin", screen="/admin/products", service="product-service PATCH /api/products/admin/{id}",
         precondition="Admin, có SP",
         steps="1. /admin/products\n2. Click 'Sửa' 1 SP\n3. Đổi stock từ 10 → 50\n4. Lưu",
         expected="DB cập nhật stock=50, audit log nếu có; UI list refresh"),
    dict(module="Admin", screen="/admin/products", service="product-service DELETE soft",
         precondition="Admin, SP không có order pending",
         steps="1. /admin/products\n2. Click 'Xóa' 1 SP\n3. Confirm",
         expected="Soft delete (deleted=true), SP biến mất khỏi /products public list"),
    dict(module="Admin", screen="/admin/orders", service="order-service GET /admin/orders",
         precondition="Admin đăng nhập, có orders trong DB",
         steps="1. /admin/orders",
         expected="Bảng tất cả orders của hệ thống (mọi user), filter status/date/user"),
    dict(module="Admin", screen="/admin/orders/[id]", service="order-service GET /admin/orders/{id}",
         precondition="Admin, có order",
         steps="1. /admin/orders\n2. Click vào 1 order",
         expected="Trang chi tiết admin (đầy đủ thông tin user + items + payment), có nút đổi status"),
    dict(module="Admin", screen="/admin/orders/[id]", service="order-service PATCH state",
         precondition="Admin, order PENDING",
         steps="1. Mở chi tiết order\n2. Đổi status → CONFIRMED\n3. Lưu",
         expected="Status order = CONFIRMED, log lưu lại, user thấy update khi xem /profile/orders/{id}"),
    dict(module="Admin", screen="/admin/coupons", service="coupon-service CRUD",
         precondition="Admin đăng nhập",
         steps="1. /admin/coupons\n2. Tạo coupon mới (code, type, value, minOrder, maxUses, expiresAt)\n3. Lưu",
         expected="Coupon mới xuất hiện trong list, có thể apply ở checkout"),
    dict(module="Admin", screen="/admin/users", service="user-service GET /admin/users",
         precondition="Admin đăng nhập",
         steps="1. /admin/users",
         expected="Danh sách users với role, status, action; có thể tìm kiếm"),
    dict(module="Admin", screen="/admin/reviews", service="review-service admin moderation",
         precondition="Admin, có reviews mới",
         steps="1. /admin/reviews\n2. Xem review pending\n3. Approve hoặc Reject",
         expected="Review chuyển trạng thái, public list /products/[slug] cập nhật"),

    # ===== CHATBOT =====
    dict(module="Chatbot", screen="Chat widget (góc dưới phải)", service="frontend /api/chat",
         precondition="Đã mở bất kỳ trang nào",
         steps="1. Click icon chat\n2. Hỏi 'Cho tôi xem các sản phẩm phở'",
         expected="Bot trả lời với danh sách SP phở (có link), không bị 500"),
    dict(module="Chatbot", screen="Chat widget", service="frontend /api/chat tool use",
         precondition="Chat widget mở, đã đăng nhập",
         steps="1. Hỏi 'Trạng thái đơn hàng gần nhất của tôi?'",
         expected="Bot gọi tool, trả về thông tin đơn hàng đúng của user hiện tại (không leak user khác)"),
    dict(module="Chatbot", screen="Chat widget", service="frontend rate limit / abuse",
         precondition="Spam 50 message liên tiếp",
         steps="1. Gửi nhiều message rapid-fire",
         expected="Rate-limit hoặc queue, không crash backend, bot trả lỗi friendly nếu vượt"),

    # ===== CROSS-CUTTING =====
    dict(module="Cross-cutting", screen="Tất cả", service="http.ts 401 handler",
         precondition="Token cũ hết hạn trong localStorage",
         steps="1. Mở trang cần auth (vd /profile)\n2. Trigger fetch (auto)",
         expected="Redirect về /login?returnTo=<currentPath>; tokens bị clear; sau khi login lại quay về đúng trang"),
    dict(module="Cross-cutting", screen="Tất cả", service="error envelope",
         precondition="Backend trả 4xx/5xx có envelope { data: { code, message } }",
         steps="Trigger 1 lỗi backend bất kỳ (vd POST cart item stock=0)",
         expected="UI hiển thị message tiếng Việt rõ ràng theo code (STOCK_SHORTAGE → 'Sản phẩm không đủ tồn kho'…)"),
    dict(module="Cross-cutting", screen="Mobile viewport", service="responsive UI",
         precondition="Browser DevTools resize ≤ 768px",
         steps="1. Mở /products, /cart, /checkout, /profile",
         expected="Layout không vỡ, menu thu thành hamburger, button đủ lớn để tap"),
    dict(module="Cross-cutting", screen="Tất cả", service="React Query cache",
         precondition="Cart có items",
         steps="1. Add 1 SP từ /products/[slug]\n2. Vào /cart ngay lập tức",
         expected="Cart hiển thị item vừa add (cache invalidated), không cần F5"),
]


# ============================================================
# COLUMN DEFINITIONS
# ============================================================

COLUMNS = [
    ("STT", 6),
    ("Module", 16),
    ("Màn hình / URL", 28),
    ("Service / API", 36),
    ("Điều kiện tiền đề", 38),
    ("Thao tác thực hiện", 50),
    ("Kết quả mong đợi", 50),
    ("Người test", 14),
    ("Ngày test", 14),
    ("Kết quả", 12),
]


# ============================================================
# STYLE HELPERS
# ============================================================

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)


def build_workbook(out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ----- Header row -----
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32

    # ----- Data rows -----
    for i, tc in enumerate(TESTCASES, start=1):
        row = i + 1  # +1 for header
        values = [
            i,
            tc["module"],
            tc["screen"],
            tc["service"],
            tc["precondition"],
            tc["steps"],
            tc["expected"],
            "",  # người test
            "",  # ngày test
            "",  # kết quả (dropdown)
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if col_idx in (1, 8, 9, 10):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = CELL_ALIGN
        # Estimate row height based on longest text
        max_lines = max(
            str(values[4]).count("\n") + len(str(values[4])) // 40 + 1,
            str(values[5]).count("\n") + len(str(values[5])) // 50 + 1,
            str(values[6]).count("\n") + len(str(values[6])) // 50 + 1,
        )
        ws.row_dimensions[row].height = max(32, min(180, 16 * max_lines))

    last_row = len(TESTCASES) + 1

    # ----- Data Validation: Kết quả dropdown -----
    dv = DataValidation(
        type="list",
        formula1='"OK,NG,NA"',
        allow_blank=True,
        showDropDown=False,  # False means dropdown IS shown (Excel inversion)
        errorStyle="warning",
        error="Chỉ chấp nhận OK, NG hoặc NA.",
        errorTitle="Giá trị không hợp lệ",
        prompt="Chọn OK / NG / NA",
        promptTitle="Kết quả test",
    )
    dv.add(f"J2:J{last_row}")
    ws.add_data_validation(dv)

    # ----- Conditional Formatting: row color theo Kết quả -----
    fill_ok = PatternFill("solid", fgColor="C6EFCE")   # green
    fill_ng = PatternFill("solid", fgColor="FFC7CE")   # red
    fill_na = PatternFill("solid", fgColor="E7E6E6")   # gray
    font_ok = Font(name="Calibri", size=10, color="006100")
    font_ng = Font(name="Calibri", size=10, color="9C0006")
    font_na = Font(name="Calibri", size=10, color="595959")

    range_str = f"A2:J{last_row}"
    ws.conditional_formatting.add(
        range_str,
        FormulaRule(formula=['$J2="OK"'], stopIfTrue=False, fill=fill_ok, font=font_ok),
    )
    ws.conditional_formatting.add(
        range_str,
        FormulaRule(formula=['$J2="NG"'], stopIfTrue=False, fill=fill_ng, font=font_ng),
    )
    ws.conditional_formatting.add(
        range_str,
        FormulaRule(formula=['$J2="NA"'], stopIfTrue=False, fill=fill_na, font=font_na),
    )

    # ----- Freeze + AutoFilter -----
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{last_row}"

    # ----- Print setup (A4 landscape) -----
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"

    # ----- Summary sheet -----
    summary = wb.create_sheet("Tổng quan")
    summary["A1"] = "Tổng số test case"
    summary["B1"] = len(TESTCASES)
    summary["A2"] = "OK"
    summary["B2"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"OK")'
    summary["A3"] = "NG"
    summary["B3"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"NG")'
    summary["A4"] = "NA"
    summary["B4"] = f'=COUNTIF(\'Test Cases\'!J2:J{last_row},"NA")'
    summary["A5"] = "Chưa test"
    summary["B5"] = f'=B1-B2-B3-B4'
    summary["A6"] = "Tỷ lệ pass (OK / Đã test)"
    summary["B6"] = f'=IFERROR(B2/(B2+B3),0)'
    summary["B6"].number_format = "0.00%"
    for r in range(1, 7):
        summary.cell(row=r, column=1).font = Font(bold=True)
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 16

    wb.save(out_path)
    print(f"[ok] wrote {out_path} ({len(TESTCASES)} test cases)")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "testcases.xlsx"
    build_workbook(out)
